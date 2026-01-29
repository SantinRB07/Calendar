from __future__ import annotations

from datetime import datetime, time
import re
from typing import Optional, Dict, List, Tuple
import asyncio
import pandas as pd
from xbbg import blp

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from prefect.utilities.asyncutils import run_sync_in_worker_thread

from tickers_earns import TICKERS_BY_SECTOR
from prefect import flow


# =========================
# CONFIG
# =========================
HIST_FIELD = "EARN_ANN_DT_TIME_HIST_WITH_EPS"

DEFAULT_TIME_MAP = {
    "AFT-MKT": time(18, 0),
    "BEF-MKT": time(8, 0),
    "DUR-MKT": time(12, 0),
}

OUTPUT_XLSX = "earnings_calendar.xlsx"

# Se True, cria uma aba por setor (pode dar muitas abas se você tiver muitos setores)
CREATE_SHEETS_PER_SECTOR = True


# =========================
# HELPERS
# =========================
def parse_time(val) -> Optional[time]:
    if pd.isna(val):
        return None
    s = str(val).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return time(h, mi)
        return None

    key = s.upper().replace(" ", "").replace("_", "-")
    key = key.replace("AFTMKT", "AFT-MKT").replace("BEFMKT", "BEF-MKT").replace("DURMKT", "DUR-MKT")
    return DEFAULT_TIME_MAP.get(key)


def short_ticker(bbg_ticker: str) -> str:
    return bbg_ticker.split()[0].strip()


def flatten_sector_dict(sector_dict: Dict[str, List[str]]) -> pd.DataFrame:
    rows = []
    for sector, tickers in sector_dict.items():
        for t in tickers:
            t = str(t).strip()
            if not t:
                continue
            rows.append({"Setor": sector, "BBG Ticker": t, "Ticker": short_ticker(t)})
    df = pd.DataFrame(rows)

    # Dedup: se um ticker aparecer em 2 setores, mantém o primeiro e marca duplicados
    if not df.empty:
        df["__key"] = df["BBG Ticker"].str.upper()
        df = df.drop_duplicates(subset="__key", keep="first").drop(columns="__key")
    return df


def fetch_future_events_for_ticker(bbg_ticker: str) -> pd.DataFrame:
    df = blp.bds(bbg_ticker, HIST_FIELD)
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.copy()
    df.columns = [c.strip().lower() for c in df.columns]

    required = {"year/period", "announcement_date", "announcement_time"}
    if not required.issubset(df.columns):
        return pd.DataFrame()

    df["announcement_date"] = pd.to_datetime(df["announcement_date"], errors="coerce")
    df = df.dropna(subset=["announcement_date"])

    today = pd.Timestamp.today().normalize()
    df = df[df["announcement_date"] >= today].copy()
    return df


# =========================
# EXCEL HELPERS
# =========================
def sanitize_sheet_name(name: str) -> str:
    # Excel: não pode []:*?/\
    s = re.sub(r"[\[\]\:\*\?\/\\]", " ", str(name)).strip()
    # limite 31 chars
    return s[:31] if len(s) > 31 else s


def write_table_sheet(wb: Workbook, title: str, df: pd.DataFrame, table_name: str):
    ws = wb.create_sheet(title=title)

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c in range(1, len(row) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F5597")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # width
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    # number formats
    headers = [c.value for c in ws[1]]
    def idx(name: str) -> Optional[int]:
        try:
            return headers.index(name) + 1
        except ValueError:
            return None

    c_date = idx("Data")
    c_time = idx("Hora (parsed)")
    c_dt = idx("DataHora (local)")

    if c_date:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_date).number_format = "yyyy-mm-dd"
    if c_time:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_time).number_format = "hh:mm"
    if c_dt:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c_dt).number_format = "yyyy-mm-dd hh:mm"

    # table
    if df.shape[0] >= 1 and df.shape[1] >= 1:
        last_col = ws.cell(row=1, column=df.shape[1]).column_letter
        last_row = ws.max_row
        tab = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    ws.row_dimensions[1].height = 20
    return ws


# =========================
# MAIN PIPELINE
# =========================
def build_calendar(sector_map: Dict[str, List[str]]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    base = flatten_sector_dict(sector_map)
    if base.empty:
        return pd.DataFrame(), pd.DataFrame([{"BBG Ticker": None, "reason": "empty_sector_dict"}])

    rows = []
    invalid = []

    for _, meta in base.iterrows():
        sector = meta["Setor"]
        bbg_ticker = meta["BBG Ticker"]
        tkr = meta["Ticker"]
        print(f"Coletando dados do {tkr}")
        
        try:
            df = fetch_future_events_for_ticker(bbg_ticker)
        except Exception as e:
            invalid.append({"Setor": sector, "BBG Ticker": bbg_ticker, "reason": f"bds_error: {type(e).__name__}: {e}"})
            continue

        if df.empty:
            invalid.append({"Setor": sector, "BBG Ticker": bbg_ticker, "reason": "no_future_events_or_empty"})
            continue

        for _, r in df.iterrows():
            per = str(r["year/period"])
            d = pd.Timestamp(r["announcement_date"]).date()
            tm_raw = r["announcement_time"]
            tm = parse_time(tm_raw)
            dt_local = datetime.combine(d, tm) if tm else None

            rows.append(
                {
                    "Setor": sector,
                    "Ticker": tkr,
                    "BBG Ticker": bbg_ticker,
                    "Período": per,
                    "Data": d,
                    "Hora (raw)": None if pd.isna(tm_raw) else str(tm_raw),
                    "Hora (parsed)": tm,
                    "DataHora (local)": dt_local,
                }
            )

    cal = pd.DataFrame(rows)
    if not cal.empty:
        cal = cal.sort_values(["Setor", "Data", "Hora (parsed)", "Ticker"], na_position="last").reset_index(drop=True)

    inv = pd.DataFrame(invalid) if invalid else pd.DataFrame(columns=["Setor", "BBG Ticker", "reason"])
    return cal, inv

def get_calendar_dates_async():
    cal, inv = build_calendar(TICKERS_BY_SECTOR)

    wb = Workbook()
    wb.remove(wb.active)

    # 1) Calendar (tudo)
    write_table_sheet(
        wb,
        "Calendar",
        cal if not cal.empty else pd.DataFrame(columns=["Setor","Ticker","BBG Ticker","Período","Data","Hora (raw)","Hora (parsed)","DataHora (local)"]),
        "CalendarTable",
    )

    if not cal.empty:
        # 2) BySector
        by_sector = cal.sort_values(["Setor", "Data", "Hora (parsed)", "Ticker"], na_position="last").reset_index(drop=True)
        write_table_sheet(wb, "BySector", by_sector, "BySectorTable")

        # 3) Next per ticker (próximo evento de cada ticker)
        cal_sorted = cal.sort_values(["Ticker", "Data", "Hora (parsed)"], na_position="last")
        next_per_ticker = cal_sorted.groupby("Ticker", as_index=False).head(1).sort_values(["Setor","Data","Ticker"])
        write_table_sheet(wb, "NextPerTicker", next_per_ticker.reset_index(drop=True), "NextPerTickerTable")

    # 5) Invalid
    write_table_sheet(wb, "Invalid", inv if not inv.empty else pd.DataFrame(columns=["Setor","BBG Ticker","reason"]), "InvalidTable")

    wb.save(OUTPUT_XLSX)
    print(f"[OK] Gerado: {OUTPUT_XLSX}")
    print(f"- Linhas no calendário: {len(cal)}")
    print(f"- Invalid/sem eventos futuros: {len(inv)}")

@flow(name="Rotina Datas de divulgação de Resultados Bloomberg", log_prints=True)
async def get_calendar_dates_flow():
    print("[EARNINGS] Iniciando coleta das datas de divulgação de resultados...")

    await run_sync_in_worker_thread(get_calendar_dates_async)

    print("[EARNINGS] Coleta finalizada com sucesso.")

if __name__ == "__main__":
    asyncio.run(get_calendar_dates_flow)
